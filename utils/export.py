#!/usr/bin/env python
# -*- coding: utf-8 -*-
""" 
__author__ = silenthz 
__date__ = 2018/10/26
"""
import datetime
import calendar

from openpyxl import Workbook
from django.forms import model_to_dict
from collections import OrderedDict
from report.models import MalfunctionData, StatisticsTop10Ne


def export_top10ne(year, month):
    begin_date = datetime.datetime(year, month, 1, 0, 0, 0)
    end_datetime = datetime.datetime(year, month, calendar.mdays[month], 23, 59, 59)
    profession_list = ['Net_4G', 'Repeater', 'Net_CDMA', 'Transmission', 'Net_Optical', 'Dynamics', 'Exchange', 'Data']
    professions = ['4G网络', '直放站', 'CDMA网络', '传输专业', '光网络专业', '动力专业', '交换专业', '数据专业']
    wb = Workbook()
    # 8个专业
    for x in range(0, 8):
        top10_qs = StatisticsTop10Ne.objects.filter(yearNum=year, monthNum=month, profession=profession_list[x]).order_by('index').values('ne')
        ne_list = []
        for i in top10_qs:
            ne_list.append(i.get('ne'))
        data_qs = MalfunctionData.objects.filter(distributeTime__range=(begin_date, end_datetime), ne__in=ne_list)
        data_dicts = list(map(lambda x: model_to_dict(x), data_qs))

        ws = wb.create_sheet(title=professions[x])
        ws.title = professions[x]
        row, col = 1, 1

        # 写表头
        head_list = ['单位', '网元名称', '专业', '故障原因', '结单信息', '故障定位', '故障标题', '派单时间',  # 8
                     '部门', '故障地市', '故障单号', '工单编号', '单据状态', '故障种类',  # 13
                     '处理时长（分钟）', '挂起时长（分钟）', '故障来源', '是否处理超时',  # 18
                     '责任部门', '类型', '原专业', '原因分类']  # 22

        for head in head_list:
            ws.cell(row=row, column=col, value=head)
            col += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['N'].width = 20
        ws.column_dimensions['Q'].width = 12
        ws.column_dimensions['T'].width = 4

        row += 1
        # col = 1
        # 设置单元格值
        for d in data_dicts:
            for key, value in d.items():
                if key == 'city':
                    ws.cell(row=row, column=1, value=str(value) if value != None else '/')
                if key == 'ne':
                    ws.cell(row=row, column=2, value=str(value) if value != None else '/')
                if key == 'profession':
                    ws.cell(row=row, column=3, value=str(value) if value != None else '/')
                if key == 'malfunctionReason':
                    ws.cell(row=row, column=4, value=str(value) if value != None else '/')
                if key == 'conclusion':
                    ws.cell(row=row, column=5, value=str(value) if value != None else '/')
                if key == 'malfunctionJudgment':
                    ws.cell(row=row, column=6, value=str(value) if value != None else '/')
                if key == 'title':
                    ws.cell(row=row, column=7, value=str(value) if value != None else '/')
                if key == 'distributeTime':
                    ws.cell(row=row, column=8, value=str(value) if value != None else '/')
                if key == 'department':
                    ws.cell(row=row, column=9, value=str(value) if value != None else '/')
                if key == 'malfunctionCity':
                    ws.cell(row=row, column=10, value=str(value) if value != None else '/')
                if key == 'receiptNumber':
                    ws.cell(row=row, column=11, value=str(value) if value != None else '/')
                if key == 'receiptSerialNumber':
                    ws.cell(row=row, column=12, value=str(value) if value != None else '/')
                if key == 'receiptStatus':
                    ws.cell(row=row, column=13, value=str(value) if value != None else '/')
                if key == 'category':
                    ws.cell(row=row, column=14, value=str(value) if value != None else '/')
                if key == 'processTime':
                    ws.cell(row=row, column=15, value=str(value) if value != None else '/')
                if key == 'hangTime':
                    ws.cell(row=row, column=16, value=str(value) if value != None else '/')
                if key == 'malfunctionSource':
                    ws.cell(row=row, column=17, value=str(value) if value != None else '/')
                if key == 'isTimeOut':
                    ws.cell(row=row, column=18, value=str(value) if value != None else '/')
                if key == 'dutyDepartment':
                    ws.cell(row=row, column=19, value=str(value) if value != None else '/')
                if key == 'type':
                    ws.cell(row=row, column=20, value=str(value) if value != None else '/')
                if key == 'originProfession':
                    ws.cell(row=row, column=21, value=str(value) if value != None else '/')
                if key == 'reasonClassification':
                    ws.cell(row=row, column=22, value=str(value) if value != None else '/')
            row += 1
        # 保存修改
    wb.remove(wb['Sheet'])
    filename = "Top10Ne-" + str(year) + '-' + str(month) + ".xlsx"
    filepath = 'utils/export_excel/' + filename
    wb.save(filepath)
    return filepath, filename
