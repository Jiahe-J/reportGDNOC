#!/usr/bin/env python
# -*- coding: utf-8 -*-
""" 
@author:silenthz 
@file: data_parser.py
@time: 2018/06/06
@报表解析工具库
"""

# 处理及时率示例
import calendar
import datetime
import xlrd
from django.core.exceptions import ObjectDoesNotExist
from django.db import IntegrityError
from django.forms import model_to_dict
from openpyxl import load_workbook

from report.models import MalfunctionData, District, StatisticsMonthlyQuality, MalfunctionLongtime, MalfunctionOnTrack, StatisticsQuarterlyQuality, \
    DistrictCity, City
from utils.fetch_Ne import *

PRD_1 = ['深圳', '广州', '东莞', '佛山', ]
district_PRD_1 = District.objects.get(pk=1)
PRD_2 = ['中山', '惠州', '江门', '珠海']
district_PRD_2 = District.objects.get(pk=2)
GD_E = ['汕头', '揭阳', '潮州', '汕尾']
district_GD_E = District.objects.get(pk=3)
GD_W = ['湛江', '茂名', '阳江', '云浮']
district_GD_W = District.objects.get(pk=4)
GD_N = ['肇庆', '清远', '阳江', '梅州', '韶关', '河源']
district_GD_N = District.objects.get(pk=5)


def deal_in_time_rate_parser():
    # 解析报表文件为元组,分组,排序
    dict = {'广州': 99, '深圳': 98, '东莞': 97, '佛山': 100, '珠海': 88, '中山': 95, '惠州': 87, '江门': 85}
    sortedTuple = sorted(dict.items(), key=lambda x: x[1], reverse=True)
    cities = []
    rate = []
    for item in sortedTuple:
        cities.append(item[0])
        rate.append(item[1])
    return cities, rate


# 按季度删除原有数据
def delete_former_data(year, quarter):
    year = int(year)
    quarter = int(quarter)
    if quarter >= 1 and quarter <= 4:
        begin_datetime = datetime.date(year, quarter, 1)
        end_datetime = datetime.date(year, quarter * 3, calendar.mdays[quarter * 3])
        MalfunctionData.objects.filter(distributeTime__range=(begin_datetime, end_datetime)).delete()


# 批量导入excel .xls数据
def parse_malfunction_data_xls(filename=None, file_contents=None, has_repeat_data=False):
    if filename:
        workbook = xlrd.open_workbook(filename)
    else:
        workbook = xlrd.open_workbook(file_contents=file_contents)
    sheet = workbook.sheet_by_index(0)
    # 数据行数
    nrows = sheet.nrows

    malfunctionList = []

    for i in range(1, nrows):
        malfunctionData = MalfunctionData()
        malfunctionData.city = sheet.cell_value(i, 1)
        malfunctionData.profession = sheet.cell_value(i, 2)
        malfunctionData.department = sheet.cell_value(i, 3)
        malfunctionData.malfunctionCity = sheet.cell_value(i, 4)
        malfunctionData.receiptNumber = sheet.cell_value(i, 5)
        malfunctionData.receiptSerialNumber = sheet.cell_value(i, 6)
        malfunctionData.receiptStatus = sheet.cell_value(i, 7)
        malfunctionData.title = sheet.cell_value(i, 8)
        malfunctionData.category = sheet.cell_value(i, 9)
        malfunctionData.distributeTime = datetime.datetime.strptime(sheet.cell_value(i, 10), '%Y-%m-%d %H:%M:%S')
        processTime = sheet.cell_value(i, 11)
        malfunctionData.processTime = processTime if processTime else 0
        hangTime = sheet.cell_value(i, 12)
        malfunctionData.hangTime = hangTime if hangTime else 0
        malfunctionData.malfunctionSource = sheet.cell_value(i, 13)
        malfunctionData.isTimeOut = sheet.cell_value(i, 14)
        malfunctionData.dutyDepartment = sheet.cell_value(i, 15)
        malfunctionData.conclusion = sheet.cell_value(i, 16)
        malfunctionData.type = sheet.cell_value(i, 17)
        malfunctionData.reasonClassification = sheet.cell_value(i, 18)
        malfunctionData.malfunctionJudgment = sheet.cell_value(i, 19)
        malfunctionData.malfunctionReason = sheet.cell_value(i, 20)
        malfunctionData.originProfession = sheet.cell_value(i, 21)
        # 解决导入重复数据,但插入速度慢
        if has_repeat_data:
            try:
                MalfunctionData.objects.get(receiptNumber=sheet.cell_value(i, 5))
                malfunctionData.save()
            except ObjectDoesNotExist:
                malfunctionList.append(malfunctionData)
        else:
            malfunctionList.append(malfunctionData)
        if (len(malfunctionList) == 500):
            try:
                MalfunctionData.objects.bulk_create(malfunctionList)
                malfunctionList = []
            except IntegrityError:
                status = '有重复数据,请选择"替换导入"方式'
                return status
    MalfunctionData.objects.bulk_create(malfunctionList)


# 批量导入excel .xlsx数据
def parse_malfunction_data_xlsx(filename=None, has_repeat_data=True):
    if filename:
        workbook = load_workbook(filename, read_only=True)
    else:
        return {'status': 'fail'}

    sheet = workbook[workbook.sheetnames[0]]
    # 数据行数
    # nrows = sheet.nrows

    malfunctionList = []
    rows = sheet.rows
    finding = True
    while finding:
        head_row = rows.__next__()
        field_list = []
        for cell in head_row:
            field_list.append(cell.value)
        if "故障单号" in field_list:
            finding = False
    nums = 0
    for row in rows:
        malfunctionData = MalfunctionData()
        malfunctionData.city = city = row[field_list.index('单位')].value
        if city in PRD_1:
            malfunctionData.district = district_PRD_1
        elif city in PRD_2:
            malfunctionData.district = district_PRD_2
        elif city in GD_E:
            malfunctionData.district = district_GD_E
        elif city in GD_W:
            malfunctionData.district = district_GD_W
        elif city in GD_N:
            malfunctionData.district = district_GD_N
        profession = row[field_list.index('专业')].value
        malfunctionData.profession = profession if profession != 'WIFI' else '无线'
        malfunctionData.department = row[field_list.index('部门')].value
        malfunctionData.malfunctionCity = row[field_list.index('故障地市')].value
        malfunctionData.receiptNumber = row[field_list.index('故障单号')].value
        print(row[field_list.index('故障单号')].value)
        malfunctionData.receiptSerialNumber = row[field_list.index('工单编号')].value
        malfunctionData.receiptStatus = row[field_list.index('单据状态')].value
        title = row[field_list.index('故障标题')].value
        if len(title) > 500:
            title = title[:500]
        if title.__contains__('条告警(含'):
            # print(title)
            pattern = re.compile(r'(\(\d+条告警)')
            index = title.index(re.search(pattern, title).group(1))
            title = title[:index]
        malfunctionData.title = title
        category = row[field_list.index('故障种类')].value
        malfunctionData.category = category
        malfunctionSource = row[field_list.index('故障来源')].value
        mtype = row[field_list.index('类型')].value
        # 由故障种类划分源专业
        if category:
            if category.__contains__('4G网络'):
                malfunctionData.originProfession = 'Net_4G'
                if malfunctionSource == '集中告警系统报故障' and mtype == '处理':
                    malfunctionData.ne = fetch_4G(title)
            elif category.__contains__('3G网络'):
                malfunctionData.originProfession = 'Net_CDMA'
                if malfunctionSource == '集中告警系统报故障' and mtype == '处理':
                    malfunctionData.ne = fetch_CDMA(title)
            elif category.__contains__('直放站'):
                malfunctionData.originProfession = 'Repeater'
                if malfunctionSource == '集中告警系统报故障' and mtype == '处理':
                    malfunctionData.ne = fetch_RPT(title)
            elif category.__contains__('本地传输网') or category.__contains__('本地光缆'):
                malfunctionData.originProfession = 'Transmission'
                if malfunctionSource == '集中告警系统报故障' and mtype == '处理':
                    malfunctionData.ne = fetch_TransmissionNetwork(title)
            elif category.__contains__('光接入网') or category.__contains__('PON'):
                malfunctionData.originProfession = 'Net_Optical'
                if malfunctionSource == '集中告警系统报故障' and mtype == '处理':
                    malfunctionData.ne = fetch_OpticalNetwork(title)
            elif category.__contains__('交换接入网') or category.__contains__('AG'):
                malfunctionData.originProfession = 'Exchange'
                if malfunctionSource == '集中告警系统报故障' and mtype == '处理':
                    malfunctionData.ne = fetch_SwitchNetwork(title)
            elif category.__contains__('数据接入网'):
                malfunctionData.originProfession = 'Data'
                if malfunctionSource == '集中告警系统报故障' and mtype == '处理':
                    malfunctionData.ne = fetch_DataNetwork(title)
            elif category.__contains__('WLAN网络'):
                malfunctionData.originProfession = 'Wifi'
                if malfunctionSource == '集中告警系统报故障' and mtype == '处理':
                    malfunctionData.ne = fetch_WIFI(title)
            if profession == '动力' and malfunctionSource == '集中告警系统报故障' and mtype == '处理':
                malfunctionData.originProfession = 'Dynamics'
                malfunctionData.ne = fetch_Dynamic(title)
        # print(str(row[field_list.index('派单时间')].value)[:19])
        malfunctionData.distributeTime = datetime.datetime.strptime(str(row[field_list.index('派单时间')].value)[:19], '%Y-%m-%d %H:%M:%S')
        processTime = row[field_list.index('处理时长（分钟）')].value
        if processTime:
            malfunctionData.distributeTime += datetime.timedelta(minutes=processTime)
        malfunctionData.processTime = processTime if processTime else 0
        hangTime = row[field_list.index('挂起时长（分钟）')].value
        malfunctionData.hangTime = hangTime if hangTime else 0
        malfunctionData.malfunctionSource = malfunctionSource
        malfunctionData.isTimeOut = row[field_list.index('是否处理超时')].value
        malfunctionData.dutyDepartment = row[field_list.index('责任部门')].value
        malfunctionData.conclusion = row[field_list.index('结单信息')].value
        malfunctionData.type = mtype
        malfunctionData.reasonClassification = row[field_list.index('原因分类')].value
        malfunctionJudgment = row[field_list.index('故障定位')].value
        malfunctionData.malfunctionJudgment = malfunctionJudgment
        if malfunctionJudgment:
            if malfunctionJudgment.__contains__('其它') or malfunctionJudgment.__contains__('其他'):
                malfunctionData.sortedReason = 1
            elif malfunctionJudgment.__contains__('动环') \
                    or malfunctionJudgment.__contains__('动力监控设备') \
                    or malfunctionJudgment.__contains__('空调') \
                    or malfunctionJudgment.__contains__('配电') \
                    or malfunctionJudgment.__contains__('电源'):
                malfunctionData.sortedReason = 5
            elif malfunctionJudgment.__contains__('数据配置'):
                malfunctionData.sortedReason = 2
            elif malfunctionJudgment.__contains__('设备故障'):
                malfunctionData.sortedReason = 3
            elif malfunctionJudgment.__contains__('线路故障'):
                malfunctionData.sortedReason = 4
        malfunctionData.malfunctionReason = row[field_list.index('故障原因')].value
        # 解决导入重复数据,但插入速度慢
        if has_repeat_data:
            s = MalfunctionData.objects.filter(receiptNumber=row[field_list.index('故障单号')].value)
            if s:
                s.update(**model_to_dict(malfunctionData))
                # print("重复数据: " + row[field_list.index('故障单号')].value)
            else:
                nums += 1
                # print("新导入行数：" + str(nums))
                malfunctionList.append(malfunctionData)
        if (len(malfunctionList) == 1000):
            try:
                MalfunctionData.objects.bulk_create(malfunctionList)
                malfunctionList = []
            except IntegrityError:
                status = '有重复数据,请选择"替换导入"方式'
                raise Exception(status)
    if malfunctionList:
        MalfunctionData.objects.bulk_create(malfunctionList)


# 解析"指标.xls"
def parse_indicators_xls(file_contents, type, year, month=None, quarter=None):
    # f = open('/Users/silenthz/Desktop/数据可视化项目/数据清单/指标.xls', 'rb')
    # file_contents = f.read()
    workbook = xlrd.open_workbook(file_contents=file_contents)
    sheet = workbook.sheet_by_index(0)
    # 数据行数
    nrows = sheet.nrows
    # result_list = []
    index = 0
    finding = True
    len = 0
    while finding:
        city = sheet.cell_value(index, 0)
        if city == '单位':
            finding = False
        index += 1

    for i in range(index, nrows):
        # print(sheet.cell_value(i, 0))
        city = sheet.cell_value(i, 0)
        if city != '??' and city != '无线' and city != '传输' and city != '动力' and city != '交换' and city != '接入网':
            if type == 'month':
                StatisticsMonthlyQuality.objects.get_or_create(
                    city=city,
                    yearNum=year,
                    monthNum=month,
                    signRate=sheet.cell_value(i, 4),
                    autoRate=sheet.cell_value(i, 11),
                    dealRate=sheet.cell_value(i, 18)
                )
            elif type == 'quarter':
                year = int(year)
                quarter = int(quarter)
                begin_date = datetime.date(year, (quarter - 1) * 3 + 1, 1)
                end_date = datetime.date(year, quarter * 3, calendar.mdays[quarter * 3])
                if city != '合计':
                    StatisticsQuarterlyQuality.objects.get_or_create(beginDate=begin_date, endDate=end_date, city=city)
                    s = StatisticsQuarterlyQuality.objects.filter(beginDate=begin_date, endDate=end_date, city=city)
                    s.update(SignRate=sheet.cell_value(i, 4))

                    pass
                pass
        pass


# 解析长历时工单.xls
def parse_malfunction_longtime(file_contents):
    # f = open('/Users/silenthz/Desktop/周报数据清单/超72小时工单.xls', 'rb')
    # file_contents = f.read()

    workbook = xlrd.open_workbook(file_contents=file_contents)
    sheet = workbook.sheet_by_index(0)
    field_list = []
    item_list = []
    for field in sheet.row(0):
        field_list.append(field.value)
    index = 0
    for i in range(1, sheet.nrows):
        title = sheet.row(i)[field_list.index('故障标题')].value
        if len(title) > 500:
            title = title[:500]
        category = sheet.row(i)[field_list.index('故障种类')].value
        city = sheet.row(i)[field_list.index('故障地市')].value
        processTime = sheet.row(i)[field_list.index('故障历时(分)')].value
        processTime = processTime if processTime else 0
        errorPosition = sheet.row(i)[field_list.index('故障位置')].value
        receiptNumber = sheet.row(i)[field_list.index('故障单号')].value
        pattern = r".*?-(.*?)-"
        year = re.search(pattern, receiptNumber).group(1)[:4]
        if title.__contains__('新开') and processTime >= 4320 and city == '广州':
            pass
        elif title.__contains__('遗留跟踪'):
            pass
        elif category.__contains__('集团NOC电子工单') or category.__contains__('网络安全') or category.__contains__('疑似光缆段故障'):
            pass
        elif processTime >= 4320 and category.__contains__('业务感知'):
            pass
        elif title.__contains__('出入') and city == '省公司':
            pass
        elif int(year) < 2017:
            pass
        else:
            if index == 0:
                # 清除旧数据
                MalfunctionLongtime.objects.all().delete()
                index += 1
            MalfunctionLongtime.objects.get_or_create(receiptNumber=receiptNumber,
                                                      title=title,
                                                      category=category,
                                                      city=city,
                                                      processTime=processTime,
                                                      errorPosition=errorPosition)
            # item.title = title
            # item.category = category
            # item.city = city
            # item.processTime = processTime
            # item.errorPosition = errorPosition
            # item_list.append(item)

            # if len(item_list) == 2000:
            #     MalfunctionLongtime.objects.bulk_create(item_list)
            #     item_list = []
    # MalfunctionLongtime.objects.bulk_create(item_list)

    # 解析"遗留跟踪单.xls"


def parse_malfunction_track(file_contents):
    """
    是否删除旧数据？
    """

    # f = open('/Users/silenthz/Desktop/周报数据清单/遗留跟踪单.xls', 'rb')
    # file_contents = f.read()
    workbook = xlrd.open_workbook(file_contents=file_contents)
    sheet = workbook.sheet_by_index(0)
    field_list = []
    index = 0
    for field in sheet.row(0):
        field_list.append(field.value)
    for i in range(1, sheet.nrows):
        title = sheet.row(i)[field_list.index('故障标题')].value
        if len(title) > 500:
            title = title[:500]
        category = sheet.row(i)[field_list.index('故障种类')].value
        receiptStatus = sheet.row(i)[field_list.index('故障状态')].value
        city = sheet.row(i)[field_list.index('故障地市')].value
        createTime = sheet.row(i)[field_list.index('建单时间')].value
        receiptNumber = sheet.row(i)[field_list.index('故障单号')].value
        if category.__contains__('业务单'):
            pass
        elif city == '省公司' and category.__contains__("IPRAN"):
            pass
        else:
            if index == 0:
                MalfunctionOnTrack.objects.all().delete()
                index += 1
            item = MalfunctionOnTrack.objects.get_or_create(receiptNumber=receiptNumber)
            item[0].title = title
            item[0].category = category
            item[0].receiptStatus = receiptStatus
            item[0].city = city
            item[0].createTime = createTime
            item[0].save()
